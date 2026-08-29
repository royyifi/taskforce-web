#!/usr/bin/env python3
"""Import the three partner sheets into a JSON file for the Prisma seed."""
import json
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT.parent / "1 Pemetaann Bidang Kerjasama.xlsx"
OUT = ROOT / "prisma" / "imported-data.json"

SHEETS = {"M. Lokal": "LOKAL", "M. Nasional ": "NASIONAL", "M. Internasional ": "INTERNASIONAL"}
FIELDS = {10: "MG", 11: "AM", 12: "RS", 13: "PI", 14: "PK", 15: "KW", 16: "KKN"}

def text(value):
    if value is None: return None
    value = str(value).strip()
    return value or None

def clean_phone(value):
    if value is None: return None
    if isinstance(value, float) and value.is_integer(): return str(int(value))
    return text(value)

def parse_dates(period):
    if not period: return None, None
    matches = re.findall(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", period)
    if not matches: return None, None
    def iso(match):
        d, m, y = map(int, match)
        try: return datetime(y, m, d).isoformat()
        except ValueError: return None
    return iso(matches[0]), iso(matches[1]) if len(matches) > 1 else None

def category(value):
    value = (text(value) or "").lower()
    if "perguruan" in value or "univers" in value or "institut" in value or "politeknik" in value: return "PERGURUAN_TINGGI"
    if "pemerintah" in value or "dinas" in value or "instansi" in value: return "PEMERINTAH"
    if "bumn" in value: return "BUMN"
    if "bumd" in value or "bumdes" in value: return "BUMD"
    if "sekolah" in value: return "SEKOLAH"
    if "ngo" in value or "yayasan" in value: return "NGO"
    if "umkm" in value or "kub" in value: return "UMKM"
    if "industri" in value or "pt " in value or value.startswith("pt.") or "cv" in value: return "INDUSTRI"
    if "swasta" in value: return "SWASTA"
    return "LAINNYA"

def slugify(value):
    value = re.sub(r"[^a-z0-9\s-]", "", value.lower())
    return re.sub(r"[-\s]+", "-", value).strip("-")

def main():
    if not XLSX.exists(): raise FileNotFoundError(XLSX)
    wb = load_workbook(XLSX, data_only=True)
    partners = []
    seen = set()
    for sheet, level in SHEETS.items():
        ws = wb[sheet]
        for row in range(5, ws.max_row + 1):
            name = text(ws.cell(row, 2).value)
            if not name or name.lower() == "nama mitra": continue
            key = re.sub(r"[^a-z0-9]", "", name.lower())
            if key in seen: continue
            seen.add(key)
            start, end = parse_dates(text(ws.cell(row, 5).value))
            fields = [code for col, code in FIELDS.items() if text(ws.cell(row, col).value)]
            partners.append({
                "name": name, "slug": slugify(name), "level": level,
                "category": category(ws.cell(row, 19).value),
                "address": text(ws.cell(row, 6).value), "phone": clean_phone(ws.cell(row, 7).value),
                "email": text(ws.cell(row, 8).value), "website": text(ws.cell(row, 9).value),
                "agreementNumber": text(ws.cell(row, 4).value), "agreementStart": start, "agreementEnd": end,
                "fields": fields, "usedByProdi": text(ws.cell(row, 17).value),
                "activityType": text(ws.cell(row, 18).value), "source": text(ws.cell(row, 20).value) or "Excel",
                "note": text(ws.cell(row, 20).value),
            })
    # Ensure unique slugs after normalization
    counts = {}
    for item in partners:
        base = item["slug"] or "mitra"
        counts[base] = counts.get(base, 0) + 1
        if counts[base] > 1: item["slug"] = f"{base}-{counts[base]}"
    OUT.write_text(json.dumps(partners, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Imported {len(partners)} partners to {OUT}")

if __name__ == "__main__": main()
